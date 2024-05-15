import xlrd
wb = xlrd.open_workbook("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/sub_variables_scores_results_by_Tongyi.xls")
sheet = wb.sheet_by_index(0)
rows = sheet.nrows
sub_scores = {}
for i in range(1, rows):
    file_path = sheet.cell(i, 0).value
    nature = sheet.cell(i, 1).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path] = {'政策类型':{ii.split(':')[0]:int(ii.split(':')[1]) for ii in nature}}
    area = sheet.cell(i, 2).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策范围'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in area}
    timeliness = sheet.cell(i, 3).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策时效性'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in timeliness}
    tools = sheet.cell(i, 4).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策工具'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in tools}
    release_agency = sheet.cell(i, 5).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策发布机构'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in release_agency}
    implementation_agency = sheet.cell(i, 6).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策执行机构'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in implementation_agency}
    function = sheet.cell(i, 7).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策功能'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in function}
    measures = sheet.cell(i, 8).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策措施'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in measures}
    coverage = sheet.cell(i, 9).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策覆盖对象'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in coverage}


# print(sub_scores.keys())
# print(sub_scores[list(sub_scores.keys())[0]])

for i in sub_scores.keys():
    pmc_index = 0
    for j in sub_scores[i].keys():
        mv_score = sum([int(sub_scores[i][j][k]) for k in sub_scores[i][j].keys()])/len(sub_scores[i][j])
        sub_scores[i][j]['主变量评分'] = mv_score
        # print(sub_scores[i][j]['主变量评分'])
        pmc_index += mv_score
    #     print(mv_score)
    # pmc_index = sum([int(sub_scores[i][j]['主变量评分']) for j in sub_scores[i].keys()])
    print(pmc_index)
    sub_scores[i]['PMC指数'] = pmc_index

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 写入表头
ws.cell(row=1, column=1, value="主变量")
# 写入第一列
row_0 = 2
for i in sub_scores[list(sub_scores.keys())[0]].keys():
    ws.cell(row=row_0, column=1, value=i)
    row_0 += 1
# ws.cell(row=row_0, column=1, value='PMC指数')


# ws['A1'] = "主变量"
column = 2
for i in sub_scores.keys():
    ws.cell(row=1, column=column, value=i)
    # 写入数值
    row_temp = 2
    temp_list = list(sub_scores[i].keys())
    temp_list.remove('PMC指数')
    for j in temp_list:
        # print(sub_scores[i][j]['主变量评分'])
        ws.cell(row=row_temp, column=column, value=sub_scores[i][j]['主变量评分'])
        row_temp += 1
    ws.cell(row=row_temp, column=column, value=sub_scores[i]['PMC指数'])
    column += 1


wb.save("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/main_variable_scores_&_PMC_index.xlsx")







