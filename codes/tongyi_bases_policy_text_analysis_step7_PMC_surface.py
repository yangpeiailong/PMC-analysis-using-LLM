import openpyxl
workbook = openpyxl.load_workbook("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/main_variable_scores_&_PMC_index.xlsx", data_only=True)
sheet = workbook.active

list_main_variables = []
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=24, max_row=10):  # 假设我们只想读取前10行和前23列
    list_main_variables.append([cell.value for cell in row])

dict_main_variables = {i:{j:list_main_variables[index_j + 1][index_i + 1] for index_j, j in enumerate([list_main_variables[i][0] for i in range(1, len(list_main_variables))])} for index_i, i in enumerate(list_main_variables[0][1:])}
# print(dict_main_variables)

from openpyxl import Workbook
import os
for i in dict_main_variables.keys():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="row_1")
    ws.cell(row=1, column=3, value="row_2")
    ws.cell(row=1, column=4, value="row_3")
    ws.cell(row=2, column=1, value="colunm_1")
    ws.cell(row=3, column=1, value="colunm_2")
    ws.cell(row=4, column=1, value="colunm_3")
    list_matrix = [dict_main_variables[i][j] for j in dict_main_variables[i].keys()]
    list_matrix2 = [list_matrix[0:3], list_matrix[3:6], list_matrix[6:]]
    for j in range(len(list_matrix2)):
        for k in range(len(list_matrix2[j])):
            ws.cell(row=j + 2, column=k + 2, value=list_matrix2[j][k])
    name = os.path.basename(i).replace('.docx','')
    wb.save(
        "D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/PMC Surface/{}.xlsx".format(name))


