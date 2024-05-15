import xlrd
wb = xlrd.open_workbook("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/sub_variables_scores_results_by_Tongyi.xls")
sheet = wb.sheet_by_index(0)
rows = sheet.nrows
sub_scores = {}
for i in range(1, rows):
    file_path = sheet.cell(i, 0).value
    nature = sheet.cell(i, 1).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path] = {'政策类型':{ii.split(':')[0]:int(ii.split(':')[1]) for ii in nature}}
    area = sheet.cell(i, 2).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策范围'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in area}
    timeliness = sheet.cell(i, 3).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策时效性'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in timeliness}
    tools = sheet.cell(i, 4).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策工具'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in tools}
    release_agency = sheet.cell(i, 5).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策发布机构'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in release_agency}
    implementation_agency = sheet.cell(i, 6).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策执行机构'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in implementation_agency}
    function = sheet.cell(i, 7).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策功能'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in function}
    measures = sheet.cell(i, 8).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策措施'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in measures}
    coverage = sheet.cell(i, 9).value.replace(' ', '').replace('{', '').replace('}', '').split(',')
    sub_scores[file_path]['政策覆盖对象'] = {ii.split(':')[0]: int(ii.split(':')[1]) for ii in coverage}


# print(sub_scores.keys())
# print(sub_scores[list(sub_scores.keys())[0]])

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 写入表头
ws['A1'] = "主变量"
ws['B1'] = "子变量"
column = 3
for i in sub_scores.keys():
    ws.cell(row=1, column=column, value=i)

    # 写入数值
    row_temp = 2
    for j in sub_scores[i].keys():
        for k in sub_scores[i][j].keys():
            ws.cell(row=row_temp, column=column, value=sub_scores[i][j][k])
            row_temp += 1
    column += 1

# 写入前两列
row2 = 2
for i in sub_scores[list(sub_scores.keys())[0]].keys():
    ws.cell(row=row2, column=1, value=i)
    for j in sub_scores[list(sub_scores.keys())[0]][i].keys():
        ws.cell(row=row2, column=2, value=j)
        row2 += 1




wb.save("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/sub_variable_scores_standard.xlsx")






