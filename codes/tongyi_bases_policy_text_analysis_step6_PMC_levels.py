
import openpyxl
workbook = openpyxl.load_workbook("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/main_variable_scores_&_PMC_index.xlsx")
sheet = workbook.active
sum = 0
for row in sheet.iter_rows(min_row=2, max_col=23, max_row=10):  # 假设我们只想读取前10行和前23列
    sum += max([cell.value for cell in row][1:])

    # for cell in row:
    #     print(cell.value)  # 打印单元格的值
PMC_index = [sheet['B11:W11'][0][i].value for i in range(len(sheet['B11:W11'][0]))]

# print(PMC_index)
poor_interval = "[{},{})".format(0, "{:.2f}".format(4/9 * sum))
acceptable_interval = "[{},{})".format("{:.2f}".format(4/9 * sum), "{:.2f}".format(2/3 * sum))
excellent_interval = "[{},{})".format("{:.2f}".format(2/3 * sum), "{:.2f}".format(8/9 * sum))
perfect_interval = "[{},{}]".format("{:.2f}".format(8/9 * sum), "{:.2f}".format(sum))


from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1, value="政策一致性级别")
ws.cell(row=1, column=2, value="poor")
ws.cell(row=1, column=3, value="acceptable")
ws.cell(row=1, column=4, value="excellent")
ws.cell(row=1, column=5, value="perfect")

ws.cell(row=2, column=1, value="PMC指数")
ws.cell(row=2, column=2, value=poor_interval)
ws.cell(row=2, column=3, value=acceptable_interval)
ws.cell(row=2, column=4, value=excellent_interval)
ws.cell(row=2, column=5, value=perfect_interval)

ws.cell(row=3, column=1, value="政策数量")
ws.cell(row=3, column=2, value=len([PMC for PMC in PMC_index if PMC >= 0 and PMC < 4/9 * sum]))
ws.cell(row=3, column=3, value=len([PMC for PMC in PMC_index if PMC >= 4/9 * sum and PMC < 2/3 * sum]))
ws.cell(row=3, column=4, value=len([PMC for PMC in PMC_index if PMC >= 2/3 * sum and PMC < 8/9 * sum]))
ws.cell(row=3, column=5, value=len([PMC for PMC in PMC_index if PMC >= 8/9 * sum and PMC <= sum]))

wb.save("D:/成都理工大学重要文件夹/Text Analysis and Evaluation of China's financial inclusion Policy based on text mining/results/Evaluation criteria of the PMC index of a policy.xlsx")